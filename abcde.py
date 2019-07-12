# -*- coding: utf-8 -*-
import re
import urllib.request
from openpyxl import load_workbook
from datetime import datetime
from bs4 import BeautifulSoup
from flask import Flask
from slack import WebClient
from slackeventsapi import SlackEventAdapter

SLACK_TOKEN = "##### SECRET #####"
SLACK_SIGNING_SECRET = "##### SECRET #####"
app = Flask(__name__)
slack_events_adaptor = SlackEventAdapter(SLACK_SIGNING_SECRET, "/listening", app)
slack_web_client = WebClient(token=SLACK_TOKEN)

def set_dict_food():
    load_wb = load_workbook("\\Users\student\PycharmProjects\\aaasaaa\FData.xlsx", data_only=True)
    load_ws = load_wb['Sheet1']
    dic = dict()
    for i in range(1, 59):
        dic[load_ws.cell(i, 2).value] = load_ws.cell(i, 1).value

    return dic
food_dict = set_dict_food()

def get_food_num(txt):
    txt = txt.replace("<@UL9LG4YTS> ", "")
    for i in food_dict.keys():
        if txt == i:
            return food_dict[txt]
    return -1

def get_month(_txt):
    if "봄" in _txt:
        month = 13
    elif "여름" in _txt:
        month = 14
    elif "가을" in _txt:
        month = 15
    elif "겨울" in _txt:
        month = 16
    elif "월" in _txt:
        for t in str(_txt).split():
            if "월" in t:
                month = int(re.sub('[^0-9]', '', t))
    else:
        month = datetime.today().month
    return month

def get_url(txt, month, num):
    if "제철음식" in txt or "계절음식" in txt:
        return "https://terms.naver.com/entry.nhn?docId=1529" + str(month + 185) + "&mobile&cid=48180&categoryId=48180"
    else:
        return "https://terms.naver.com/entry.nhn?docId=" + str(num) + "&mobile&cid=48180&categoryId=48248"

def get_title(month):
        if month < 13:
            return "##### %d"%month + "월 제철음식 #####"
        elif month == 13:
            return "##### 봄(3월, 4월, 5월) 제철음식 #####"
        elif month == 14:
            return "##### 여름(6월, 7월, 8월) 계절음식 #####"
        elif month == 15:
            return "##### 가을(9월, 10월, 11월) 계절음식 #####"
        elif month == 16:
            return "##### 겨울(12월, 1월, 2월) 계절음식 #####"
        else:
            return "Error"

def adjust_food_list(_list):

    _list = str(_list[0]).split("\n")
    re_text = str(_list[1]).split("·")
    _list[1] = re_text[0]
    _list.insert(2, "·" + re_text[1])
    del _list[len(_list) - 2]
    del _list[len(_list) - 2]
    return _list

def _crawl_main(text):
    if text == "<@UL9LG4YTS>" or text.replace("<@UL9LG4YTS> ", "").lower() == "hello" or text.replace("<@UL9LG4YTS> ", "").lower() == "hi":
        return "안녕하세요! 계절음식추천 봇이에요~ \n 다음과 같이 사용해 주세요 \n\n\n @<봇이름> ()월/(계절) 제철음식/계절음식 \n @<봇이름> 추천 재료 정보 \n\n\n!help 또는 !양식"
    if "!help" in text or "!양식" in text:
        return "`@<봇이름> ()월/(계절) 제철음식/계절음식`\n `ex) @<봇이름> 10월 제철음식` \n `@<봇이름> 여름 계절음식`"

    f_num = int()
    month = get_month(text)
    if not "제철음식" in text and not "계절음식" in text:
        f_num = get_food_num(text)
        if f_num < 0:
            return "`잘못된 입력 같아요. ` \n `!help or !양식`"

    url = get_url(text, month, f_num)
    req = urllib.request.Request(url)
    sourcecode = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(sourcecode, "html.parser")

    f_list = []
    if "제철음식" in text or "계절음식" in text:
        f_list.append(get_title(month))
        for food in soup.find_all("strong", class_="c-title"):
            f_list.append(food.get_text())
    else:
        data2 = []
        for data in soup.find_all('div', class_='section_wrap'):
            name = data.find('h2', class_='headword').get_text()
            for d in data.find_all('p', class_='txt'):
                data2.append(d.get_text())
        data2 = "".join(data2)
        f_list.append("### " + name + " ###\n" + data2)
        f_list = adjust_food_list(f_list)

    return '\n'.join(f_list)

@slack_events_adaptor.on("app_mention")
def app_mentioned(event_data):
    channel = event_data["event"]["channel"]
    text = event_data["event"]["text"]

    message = _crawl_main(text)
    slack_web_client.chat_postMessage(
        channel=channel,
        text=message
    )

# / 로 접속하면 서버가 준비되었다고 알려줍니다.
@app.route("/", methods=["GET"])
def index():
    return "<h1>Server is ready.</h1>"

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=4040)

